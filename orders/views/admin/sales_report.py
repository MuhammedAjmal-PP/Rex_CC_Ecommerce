"""
Admin sales report views — page, PDF download, Excel download.
"""

import io
from datetime import datetime
from decimal import Decimal

from django.contrib.auth.decorators import user_passes_test
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import render
from django.template.loader import render_to_string
from django.utils import timezone
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_GET

from orders.service.sales_report import get_date_range, get_sales_report
from weasyprint import HTML


# ────────────────────────────────────────────
# Helpers
# ────────────────────────────────────────────


def _parse_filters(request):
    """
    Extract and validate filter params from the GET querystring.
    Returns (filter_type, start_dt, end_dt, label).
    """
    filter_type = request.GET.get("filter", "1_day").strip()
    start_str = request.GET.get("start_date", "").strip()
    end_str = request.GET.get("end_date", "").strip()

    start_date = end_date = None
    if filter_type == "custom" and start_str and end_str:
        try:
            start_date = timezone.make_aware(
                datetime.strptime(start_str, "%Y-%m-%d")
            )
            end_date = timezone.make_aware(
                datetime.strptime(end_str, "%Y-%m-%d").replace(
                    hour=23, minute=59, second=59
                )
            )
        except ValueError:
            filter_type = "1_day"

    start_dt, end_dt = get_date_range(filter_type, start_date, end_date)

    # Human-readable label
    labels = {
        "1_day": "Today",
        "1_week": "Last 7 Days",
        "1_month": "Last 30 Days",
        "custom": f"{start_dt.strftime('%b %d, %Y')} – {end_dt.strftime('%b %d, %Y')}",
    }
    label = labels.get(filter_type, "Today")

    return filter_type, start_dt, end_dt, label


def _build_filter_qs(request):
    """Build the query string for download links."""
    params = []
    for key in ("filter", "start_date", "end_date"):
        val = request.GET.get(key)
        if val:
            params.append(f"{key}={val}")
    return "&".join(params)


# ────────────────────────────────────────────
# Report Page
# ────────────────────────────────────────────


@user_passes_test(lambda u: u.is_superuser, login_url="admin_login")
@never_cache
@require_GET
def sales_report_view(request):
    filter_type, start_dt, end_dt, label = _parse_filters(request)
    report = get_sales_report(start_dt, end_dt)

    paginator = Paginator(report["orders_qs"], 15)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    context = {
        "filter_type": filter_type,
        "start_date": request.GET.get("start_date", ""),
        "end_date": request.GET.get("end_date", ""),
        "period_label": label,
        "total_orders": report["total_orders"],
        "total_order_amount": report["total_order_amount"],
        "total_discount": report["total_discount"],
        "total_coupon_discount": report["total_coupon_discount"],
        "orders": page_obj,
        "filter_qs": _build_filter_qs(request),
    }
    return render(request, "orders/admin/sales_report.html", context)


# ────────────────────────────────────────────
# PDF Download
# ────────────────────────────────────────────


@user_passes_test(lambda u: u.is_superuser, login_url="admin_login")
@never_cache
@require_GET
def download_sales_report_pdf(request):
    filter_type, start_dt, end_dt, label = _parse_filters(request)
    report = get_sales_report(start_dt, end_dt)
    orders = list(report["orders_qs"])  # all results — no pagination

    html_string = render_to_string(
        "orders/admin/sales_report_pdf.html",
        {
            "period_label": label,
            "start_date": start_dt,
            "end_date": end_dt,
            "total_orders": report["total_orders"],
            "total_order_amount": report["total_order_amount"],
            "total_discount": report["total_discount"],
            "total_coupon_discount": report["total_coupon_discount"],
            "orders": orders,
        },
    )

    pdf = HTML(string=html_string).write_pdf()
    filename = f"sales_report_{start_dt.strftime('%Y%m%d')}_{end_dt.strftime('%Y%m%d')}.pdf"

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ────────────────────────────────────────────
# Excel Download
# ────────────────────────────────────────────


@user_passes_test(lambda u: u.is_superuser, login_url="admin_login")
@never_cache
@require_GET
def download_sales_report_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    filter_type, start_dt, end_dt, label = _parse_filters(request)
    report = get_sales_report(start_dt, end_dt)
    orders = report["orders_qs"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # ── Styles ──
    header_font = Font(name="Calibri", bold=True, size=14)
    sub_font = Font(name="Calibri", size=11, color="555555")
    col_header_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    col_header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    col_header_alignment = Alignment(horizontal="center", vertical="center")
    currency_fmt = '#,##0.00'
    thin_border = Border(
        bottom=Side(style="thin", color="DDDDDD"),
    )

    # ── Title ──
    ws.merge_cells("A1:G1")
    ws["A1"] = "REX CC — Sales Report"
    ws["A1"].font = header_font

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Period: {label}"
    ws["A2"].font = sub_font

    # ── Summary ──
    ws.append([])  # row 3 blank
    summary_data = [
        ("Total Orders", report["total_orders"]),
        ("Total Order Amount (₹)", report["total_order_amount"]),
        ("Total Discount (₹)", report["total_discount"]),
        ("Total Coupon Discount (₹)", report["total_coupon_discount"]),
    ]
    summary_label_font = Font(name="Calibri", bold=True, size=10)
    for label_text, value in summary_data:
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=label_text).font = summary_label_font
        cell = ws.cell(row=row, column=2, value=float(value) if isinstance(value, Decimal) else value)
        if isinstance(value, Decimal):
            cell.number_format = currency_fmt

    ws.append([])  # spacer

    # ── Column Headers ──
    columns = [
        ("Order #", 18),
        ("Date", 18),
        ("Customer", 28),
        ("Payment", 14),
        ("Discount (₹)", 16),
        ("Coupon (₹)", 16),
        ("Grand Total (₹)", 18),
    ]
    header_row = ws.max_row + 1
    for col_idx, (col_name, width) in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = col_header_alignment
        ws.column_dimensions[cell.column_letter].width = width

    # ── Data Rows ──
    data_font = Font(name="Calibri", size=10)
    for order in orders:
        txn = order.payment_transaction
        payment_label = txn.get_payment_method_display() if txn else "—"
        customer = "—"
        if order.user:
            name = f"{order.user.first_name or ''} {order.user.last_name or ''}".strip()
            customer = name if name else order.user.email

        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=order.order_number).font = data_font
        ws.cell(row=row, column=2, value=order.created_at.strftime("%Y-%m-%d %I:%M %p")).font = data_font
        ws.cell(row=row, column=3, value=customer).font = data_font
        ws.cell(row=row, column=4, value=payment_label).font = data_font

        discount_cell = ws.cell(row=row, column=5, value=float(order.discount))
        discount_cell.number_format = currency_fmt
        discount_cell.font = data_font

        coupon_cell = ws.cell(row=row, column=6, value=float(order.coupon_discount))
        coupon_cell.number_format = currency_fmt
        coupon_cell.font = data_font

        total_cell = ws.cell(row=row, column=7, value=float(order.grand_total))
        total_cell.number_format = currency_fmt
        total_cell.font = data_font

        # light bottom border
        for c in range(1, 8):
            ws.cell(row=row, column=c).border = thin_border

    # ── Write to response ──
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"sales_report_{start_dt.strftime('%Y%m%d')}_{end_dt.strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
